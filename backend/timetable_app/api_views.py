"""
RESTful API endpoints for timetable management.

Provides CRUD operations on all entities and specialized operations
for data import, timetable generation, and export.
"""

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.contrib.auth import authenticate
from django.http import HttpResponse
from datetime import datetime
import io
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import (
    User, School, Department, Course, Lecturer, Room,
    TimeSlot, LecturerAvailability, TimetableEntry,
    CommonUnit, RecurrentUnit, TrainingData
)
from .serializers import (
    UserSerializer, RegisterSerializer, LoginSerializer,
    SchoolSerializer, DepartmentSerializer, CourseSerializer,
    LecturerSerializer, RoomSerializer, TimeSlotSerializer,
    LecturerAvailabilitySerializer, TimetableEntrySerializer,
    CommonUnitSerializer, RecurrentUnitSerializer, TrainingDataSerializer
)

from .services import TimetableScheduler, FileParser


# ============================================================================
# Custom Permissions
# ============================================================================

class IsTimetablerUser(permissions.BasePermission):
    """Allow access only to timetabler (admin) users."""
    
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and request.user.is_timetabler
        )


class IsLecturerUser(permissions.BasePermission):
    """Allow access to lecturer and timetabler users."""
    
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and (request.user.is_lecturer or request.user.is_timetabler)
        )


class IsTimetablerOrReadOnly(permissions.BasePermission):
    """
    Allow full CRUD for timetablers; read-only for lecturers and students.
    """
    
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user.is_authenticated
        return (
            request.user.is_authenticated
            and request.user.is_timetabler
        )


class CanEditTimetable(permissions.BasePermission):
    """
    Allow timetablers to create/edit timetables.
    Lecturers and students are read-only.
    """
    
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user.is_authenticated
        return (
            request.user.is_authenticated
            and request.user.is_timetabler
        )


class IsAdminUserOrReadOnly(permissions.BasePermission):
    """
    Backward compatibility: Allow full CRUD for timetablers;
    read-only for other users.
    """
    
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user.is_authenticated
        return (
            request.user.is_authenticated
            and request.user.is_timetabler
        )


class IsStudentUser(permissions.BasePermission):
    """Allow access only to student users."""
    
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and request.user.is_student
        )


class CanAccessTrainingData(permissions.BasePermission):
    """
    Allow access to training data only to timetablers (admins).
    """
    
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and request.user.is_timetabler
        )


class CanManageLecturerData(permissions.BasePermission):
    """
    Allow lecturers to manage only relevant data:
    - View and upload courses they teach
    - Manage rooms and time slots
    - Cannot access training data
    """
    
    def has_permission(self, request, view):
        return (
            request.user.is_authenticated
            and (request.user.is_lecturer or request.user.is_timetabler)
        )


class IsStudentReadOnly(permissions.BasePermission):
    """
    Allow students read-only access to timetables.
    Students cannot view training data.
    """
    
    def has_permission(self, request, view):
        if request.method not in permissions.SAFE_METHODS:
            return False
        return request.user.is_authenticated


# ============================================================================
# Authentication Endpoints
# ============================================================================

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_csrf_token(request):
    """
    Get CSRF token for cross-origin requests.
    The CSRF token is automatically set in response cookies.
    """
    from django.middleware.csrf import get_token
    csrf_token = get_token(request)
    return Response(
        {
            'csrfToken': csrf_token,
            'message': 'CSRF token generated'
        },
        status=status.HTTP_200_OK
    )


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register_user(request):
    """
    Register a new user.
    Role is determined automatically based on email domain:
    - @admin.cuk.ac.ke -> Timetabler
    - @staff.cuk.ac.ke -> Lecturer
    - @student.cuk.ac.ke -> Student
    """
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response(
            {
                'message': 'User registered successfully',
                'user': UserSerializer(user).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_user(request):
    """
    Authenticate user and return user info.
    Only users from CUK domains can log in.
    """
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        return Response(
            {
                'message': 'Login successful',
                'user': UserSerializer(user).data
            },
            status=status.HTTP_200_OK
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def current_user(request):
    """Get current authenticated user's info."""
    return Response(UserSerializer(request.user).data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def logout_user(request):
    """Logout user (session-based)."""
    from django.contrib.auth import logout
    logout(request)
    return Response(
        {'message': 'Logged out successfully'},
        status=status.HTTP_200_OK
    )


@api_view(['PUT', 'PATCH'])
@permission_classes([permissions.IsAuthenticated])
def update_user_profile(request):
    """
    Update user profile.
    Email cannot be changed (role is determined by email).
    """
    user = request.user
    
    # Don't allow email changes as role is based on email domain
    if 'email' in request.data:
        return Response(
            {'error': 'Email cannot be changed.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if 'first_name' in request.data:
        user.first_name = request.data['first_name']
    if 'last_name' in request.data:
        user.last_name = request.data['last_name']
    
    if 'new_password' in request.data:
        current_password = request.data.get('current_password')
        if not current_password or not user.check_password(current_password):
            return Response(
                {'error': 'Current password is incorrect.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        user.set_password(request.data['new_password'])
    
    user.save()
    return Response(
        {
            'message': 'Profile updated successfully',
            'user': UserSerializer(user).data
        },
        status=status.HTTP_200_OK
    )


# ============================================================================
# CRUD ViewSets (Core Models)
# ============================================================================

class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.prefetch_related('departments').all()
    serializer_class = SchoolSerializer
    search_fields = ['name', 'code']
    
    def get_permissions(self):
        """
        Custom permission:
        - GET (list/retrieve): Allow all authenticated users
        - POST/PUT/DELETE: Allow only timetablers
        """
        if self.request.method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsTimetablerUser()]


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    search_fields = ['name', 'code']
    
    def get_permissions(self):
        """
        Custom permission:
        - GET (list/retrieve): Allow all authenticated users
        - POST/PUT/DELETE: Allow only timetablers
        """
        if self.request.method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsTimetablerUser()]


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related('department').all()
    serializer_class = CourseSerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['department', 'year_of_study', 'study_mode']
    search_fields = ['name', 'code']


class LecturerViewSet(viewsets.ModelViewSet):
    queryset = Lecturer.objects.select_related('department').all()
    serializer_class = LecturerSerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['department']
    search_fields = ['name', 'employee_id']


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['room_type', 'building']
    search_fields = ['name', 'building']


class TimeSlotViewSet(viewsets.ModelViewSet):
    queryset = TimeSlot.objects.all()
    serializer_class = TimeSlotSerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['day']


class LecturerAvailabilityViewSet(viewsets.ModelViewSet):
    queryset = LecturerAvailability.objects.select_related(
        'lecturer', 'time_slot'
    ).all()
    serializer_class = LecturerAvailabilitySerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['lecturer', 'is_available']


class TimetableEntryViewSet(viewsets.ModelViewSet):
    queryset = TimetableEntry.objects.select_related(
        'course', 'lecturer', 'room', 'time_slot'
    ).all()
    serializer_class = TimetableEntrySerializer
    permission_classes = [permissions.IsAuthenticated, IsTimetablerOrReadOnly]
    filterset_fields = ['time_slot__day', 'is_locked', 'course', 'lecturer', 'room']


# ============================================================================
# Data Import/Export Endpoints
# ============================================================================

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminUserOrReadOnly])
def upload_data(request):
    """Upload Excel or PDF file to import timetable data."""
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    success, message = FileParser.parse_file(request.FILES['file'])
    
    if not success:
        return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response(
        {'message': message},
        status=status.HTTP_201_CREATED
    )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_excel_template(request):
    """Download empty Excel template for data import."""
    wb = openpyxl.Workbook()
    
    # Departments
    ws = wb.active
    ws.title = "Departments"
    ws['A1'] = "Department Name"
    ws['B1'] = "Department Code"
    ws['A2'] = "Computer Science"
    ws['B2'] = "CS"
    
    # Lecturers
    ws = wb.create_sheet("Lecturers")
    ws['A1'] = "Lecturer Name"
    ws['B1'] = "Employee ID"
    ws['C1'] = "Department Code"
    ws['D1'] = "Email"
    ws['A2'] = "Dr. John Smith"
    ws['B2'] = "LEC001"
    ws['C2'] = "CS"
    ws['D2'] = "john@university.edu"
    
    # Rooms
    ws = wb.create_sheet("Rooms")
    ws['A1'] = "Room Name"
    ws['B1'] = "Capacity"
    ws['C1'] = "Room Type"
    ws['D1'] = "Building"
    ws['A2'] = "A101"
    ws['B2'] = 50
    ws['C2'] = "CLASSROOM"
    ws['D2'] = "Main Building"
    
    # Time Slots
    ws = wb.create_sheet("TimeSlots")
    ws['A1'] = "Day"
    ws['B1'] = "Start Time"
    ws['C1'] = "End Time"
    ws['A2'] = "Monday"
    ws['B2'] = "09:00"
    ws['C2'] = "10:00"
    
    # Courses
    ws = wb.create_sheet("Courses")
    ws['A1'] = "Course Code"
    ws['B1'] = "Course Name"
    ws['C1'] = "Department Code"
    ws['D1'] = "Year"
    ws['E1'] = "Study Mode"
    ws['F1'] = "Class Size"
    ws['G1'] = "Hours/Week"
    ws['A2'] = "CS101"
    ws['B2'] = "Intro to Programming"
    ws['C2'] = "CS"
    ws['D2'] = 1
    ws['E2'] = "IN_PERSON"
    ws['F2'] = 30
    ws['G2'] = 2
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="timetable_template.xlsx"'
    return response


# ============================================================================
# Timetable Generation & Export
# ============================================================================

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminUserOrReadOnly])
def generate_timetable(request):
    """Generate optimized timetable from existing data."""
    try:
        # Clear old entries
        TimetableEntry.objects.all().delete()
        
        # Get data
        courses = list(Course.objects.all())
        lecturers = list(Lecturer.objects.all())
        rooms = list(Room.objects.all())
        time_slots = list(TimeSlot.objects.all())
        
        if not all([courses, lecturers, rooms, time_slots]):
            return Response(
                {'error': 'Insufficient data for scheduling'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate schedule
        scheduler = TimetableScheduler(courses, lecturers, rooms, time_slots)
        entries = scheduler.generate()
        
        if not entries:
            return Response(
                {'error': 'No feasible schedule found', 'count': 0},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response(
            {'message': f'Scheduled {len(entries)} classes', 'count': len(entries)},
            status=status.HTTP_201_CREATED
        )
    
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_timetable_pdf(request):
    """Export timetable as PDF (optionally filtered)."""
    try:
        # Get filter parameters
        lecturer_id = request.query_params.get('lecturer_id')
        room_id = request.query_params.get('room_id')
        course_id = request.query_params.get('course_id')
        
        # Build query
        query = TimetableEntry.objects.select_related(
            'course', 'lecturer', 'room', 'time_slot'
        )
        
        if lecturer_id:
            query = query.filter(lecturer_id=lecturer_id)
        if room_id:
            query = query.filter(room_id=room_id)
        if course_id:
            query = query.filter(course_id=course_id)
        
        entries = list(query.order_by('time_slot__day', 'time_slot__start_time'))
        
        if not entries:
            return Response(
                {'error': 'No timetable entries found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(letter),
            topMargin=0.5*inch, bottomMargin=0.5*inch
        )
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=18, textColor=colors.HexColor('#1f4788'),
            spaceAfter=12
        )
        
        if lecturer_id:
            from django.shortcuts import get_object_or_404
            lecturer = get_object_or_404(Lecturer, id=lecturer_id)
            title = f"Timetable: {lecturer.name}"
        elif room_id:
            from django.shortcuts import get_object_or_404
            room = get_object_or_404(Room, id=room_id)
            title = f"Timetable: {room.name}"
        elif course_id:
            from django.shortcuts import get_object_or_404
            course = get_object_or_404(Course, id=course_id)
            title = f"Timetable: {course.code}"
        else:
            title = "Complete Timetable"
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3*inch))
        
        date_style = ParagraphStyle(
            'DateStyle', parent=styles['Normal'],
            fontSize=10, textColor=colors.grey
        )
        story.append(
            Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style)
        )
        story.append(Spacer(1, 0.2*inch))
        
        # Table data
        data = [['Day', 'Time', 'Course', 'Lecturer', 'Room', 'Size']]
        
        for entry in entries:
            day = entry.time_slot.get_day_display()
            time = f"{entry.time_slot.start_time.strftime('%H:%M')}-{entry.time_slot.end_time.strftime('%H:%M')}"
            course = entry.course.code
            lecturer = entry.lecturer.name
            room = entry.room.name
            size = entry.course.class_size
            
            data.append([day, time, course, lecturer, room, str(size)])
        
        # Create table
        table = Table(data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================================
# Filtered Timetable Views
# ============================================================================

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def course_timetable(request, course_id):
    """Get timetable for specific course."""
    entries = TimetableEntry.objects.filter(
        course_id=course_id
    ).select_related('time_slot', 'room', 'lecturer')
    return Response(TimetableEntrySerializer(entries, many=True).data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def lecturer_timetable(request, lecturer_id):
    """Get timetable for specific lecturer."""
    entries = TimetableEntry.objects.filter(
        lecturer_id=lecturer_id
    ).select_related('time_slot', 'room', 'course')
    return Response(TimetableEntrySerializer(entries, many=True).data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def room_timetable(request, room_id):
    """Get timetable for specific room."""
    entries = TimetableEntry.objects.filter(
        room_id=room_id
    ).select_related('time_slot', 'course', 'lecturer')
    return Response(TimetableEntrySerializer(entries, many=True).data)

# ============================================================================
# Training Data Endpoints
# ============================================================================

@api_view(['GET', 'POST'])
@permission_classes([CanAccessTrainingData])
def training_data_list(request):
    """
    List all training datasets (admin only).
    GET: Return all training datasets.
    POST: Create a new training dataset.
    """
    if request.method == 'GET':
        datasets = TrainingData.objects.all()
        serializer = TrainingDataSerializer(datasets, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = TrainingDataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([CanAccessTrainingData])
def training_data_detail(request, pk):
    """
    Get, update, or delete a specific training dataset (admin only).
    """
    try:
        dataset = TrainingData.objects.get(pk=pk)
    except TrainingData.DoesNotExist:
        return Response(
            {'error': 'Training dataset not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        serializer = TrainingDataSerializer(dataset)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = TrainingDataSerializer(dataset, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        dataset.delete()
        return Response(
            {'message': 'Training dataset deleted'},
            status=status.HTTP_204_NO_CONTENT
        )


@api_view(['GET', 'POST'])
@permission_classes([CanAccessTrainingData])
def common_units_list(request):
    """
    List all common units or create a new one (admin only).
    Common units are cross-departmental teaching units.
    """
    if request.method == 'GET':
        units = CommonUnit.objects.all()
        serializer = CommonUnitSerializer(units, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = CommonUnitSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([CanAccessTrainingData])
def common_unit_detail(request, pk):
    """Get, update, or delete a specific common unit (admin only)."""
    try:
        unit = CommonUnit.objects.get(pk=pk)
    except CommonUnit.DoesNotExist:
        return Response(
            {'error': 'Common unit not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        serializer = CommonUnitSerializer(unit)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = CommonUnitSerializer(unit, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        unit.delete()
        return Response(
            {'message': 'Common unit deleted'},
            status=status.HTTP_204_NO_CONTENT
        )


@api_view(['GET', 'POST'])
@permission_classes([CanAccessTrainingData])
def recurrent_units_list(request):
    """
    List all recurrent units or create a new one (admin only).
    Recurrent units are units taught in specific courses and years.
    """
    if request.method == 'GET':
        units = RecurrentUnit.objects.all()
        serializer = RecurrentUnitSerializer(units, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = RecurrentUnitSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([CanAccessTrainingData])
def recurrent_unit_detail(request, pk):
    """Get, update, or delete a specific recurrent unit (admin only)."""
    try:
        unit = RecurrentUnit.objects.get(pk=pk)
    except RecurrentUnit.DoesNotExist:
        return Response(
            {'error': 'Recurrent unit not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        serializer = RecurrentUnitSerializer(unit)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = RecurrentUnitSerializer(unit, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        unit.delete()
        return Response(
            {'message': 'Recurrent unit deleted'},
            status=status.HTTP_204_NO_CONTENT
        )


# ============================================================================
# File Upload & Data Import Endpoints
# ============================================================================

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsTimetablerUser])
def upload_training_data(request):
    """
    Upload and parse training data files (Excel/CSV/PDF).
    Creates or updates training dataset with imported data.
    
    Files should contain:
    - Courses with department info
    - Lecturers with departments
    - Rooms
    - Time slots
    """
    from .serializers import FileUploadSerializer
    
    serializer = FileUploadSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    uploaded_file = serializer.validated_data['file']
    file_type = serializer.validated_data['file_type']
    description = serializer.validated_data.get('description', '')
    
    try:
        # Parse the file
        if file_type == 'training_data':
            success, message, data_summary = FileParser.parse_training_data_file(uploaded_file)
        else:  # timetable_data
            success, message, data_summary = FileParser.parse_timetable_data_file(uploaded_file)
        
        if not success:
            return Response(
                {'error': message},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create training dataset record
        if file_type == 'training_data':
            training_data = TrainingData.objects.create(
                name=f"Training Data - {uploaded_file.name}",
                description=description or f"Imported from {uploaded_file.name}",
                created_by=request.user,
                is_active=True
            )
            
            # Add relationships based on parsed data
            if 'courses' in data_summary:
                training_data.courses.set(data_summary['courses'])
            if 'lecturers' in data_summary:
                training_data.lecturers.set(data_summary['lecturers'])
            if 'departments' in data_summary:
                training_data.departments.set(data_summary['departments'])
            
            training_data.save()
        
        return Response({
            'message': message,
            'data_summary': data_summary,
            'training_data_id': training_data.id if file_type == 'training_data' else None
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response(
            {'error': f'File processing error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsTimetablerUser])
def upload_data_file(request):
    """
    General file upload endpoint for importing timetable data.
    Supports Excel and PDF formats.
    """
    from .serializers import FileUploadSerializer
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_file = request.FILES['file']
    
    try:
        success, message = FileParser.parse_file(uploaded_file)
        
        if not success:
            return Response(
                {'error': message},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            'message': message,
            'filename': uploaded_file.name
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response(
            {'error': f'File processing error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )