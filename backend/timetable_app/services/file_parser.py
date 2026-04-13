"""
File parsing service for Excel and PDF data import.

Handles extraction of timetable data from various file formats
and populates database models.
"""

import openpyxl
import pdfplumber
from django.shortcuts import get_object_or_404
from ..models import Department, Lecturer, Room, TimeSlot, Course, School, TimetableEntry


class FileParser:
    """Parses Excel and PDF files to extract and import timetable data."""
    
    @staticmethod
    def parse_file(file):
        """
        Auto-detect file type and parse accordingly.
        
        Args:
            file: Uploaded file object
            
        Returns:
            tuple: (success: bool, message: str)
        """
        filename = file.name.lower()
        
        if filename.endswith('.xlsx'):
            return FileParser.parse_excel(file)
        elif filename.endswith('.pdf'):
            return FileParser.parse_pdf(file)
        else:
            return False, "Unsupported file type. Use .xlsx or .pdf"
    
    @staticmethod
    def parse_excel(file):
        """Parse Excel file with standard sheets."""
        try:
            wb = openpyxl.load_workbook(file)
            
            # Parse in dependency order
            FileParser._parse_departments(wb)
            FileParser._parse_lecturers(wb)
            FileParser._parse_rooms(wb)
            FileParser._parse_timeslots(wb)
            FileParser._parse_courses(wb)
            
            return True, "Excel data imported successfully"
        
        except Exception as e:
            return False, f"Error parsing Excel: {str(e)}"
    
    @staticmethod
    def parse_pdf(file):
        """Parse PDF file to extract tabular data."""
        try:
            with pdfplumber.open(file) as pdf:
                text_content = ""
                tables_data = []
                
                # Extract text and tables from all pages
                for page in pdf.pages:
                    text_content += page.extract_text() or ""
                    text_content += "\n"
                    
                    tables = page.extract_tables()
                    if tables:
                        tables_data.extend(tables)
                
                # Try parsing from tables first (more reliable)
                if tables_data:
                    FileParser._parse_tables(tables_data)
                else:
                    # Fall back to text parsing
                    FileParser._parse_text_sections(text_content)
                
                return True, "PDF data imported successfully"
        
        except Exception as e:
            return False, f"Error parsing PDF: {str(e)}"
    
    @staticmethod
    def _parse_departments(wb):
        """Extract departments from workbook."""
        if 'Departments' not in wb.sheetnames:
            return
        
        sheet = wb['Departments']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                Department.objects.get_or_create(
                    code=str(row[1]),
                    defaults={'name': str(row[0])}
                )
    
    @staticmethod
    def _parse_lecturers(wb):
        """Extract lecturers from workbook."""
        if 'Lecturers' not in wb.sheetnames:
            return
        
        sheet = wb['Lecturers']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                try:
                    dept = Department.objects.get(code=str(row[2]))
                    Lecturer.objects.get_or_create(
                        employee_id=str(row[1]),
                        defaults={
                            'name': str(row[0]),
                            'email': str(row[3]) if len(row) > 3 and row[3] else '',
                            'department': dept
                        }
                    )
                except Department.DoesNotExist:
                    pass
    
    @staticmethod
    def _parse_rooms(wb):
        """Extract rooms from workbook."""
        if 'Rooms' not in wb.sheetnames:
            return
        
        sheet = wb['Rooms']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                try:
                    capacity = int(row[1])
                    Room.objects.get_or_create(
                        name=str(row[0]),
                        defaults={
                            'capacity': capacity,
                            'room_type': str(row[2]) if len(row) > 2 and row[2] else 'CLASSROOM',
                            'building': str(row[3]) if len(row) > 3 and row[3] else ''
                        }
                    )
                except (ValueError, TypeError):
                    pass
    
    @staticmethod
    def _parse_timeslots(wb):
        """Extract time slots from workbook."""
        if 'TimeSlots' not in wb.sheetnames:
            return
        
        sheet = wb['TimeSlots']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                try:
                    TimeSlot.objects.get_or_create(
                        day=str(row[0]),
                        start_time=row[1],
                        end_time=row[2]
                    )
                except Exception:
                    pass
    
    @staticmethod
    def _parse_courses(wb):
        """Extract courses from workbook."""
        if 'Courses' not in wb.sheetnames:
            return
        
        sheet = wb['Courses']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                try:
                    dept = Department.objects.get(code=str(row[2]))
                    Course.objects.get_or_create(
                        code=str(row[0]),
                        defaults={
                            'name': str(row[1]),
                            'department': dept,
                            'year_of_study': int(row[3]) if len(row) > 3 and row[3] else 1,
                            'study_mode': str(row[4]) if len(row) > 4 and row[4] else 'IN_PERSON',
                            'class_size': int(row[5]) if len(row) > 5 and row[5] else 30,
                            'hours_per_week': float(row[6]) if len(row) > 6 and row[6] else 1.0
                        }
                    )
                except (Department.DoesNotExist, ValueError, TypeError):
                    pass
    
    @staticmethod
    def _parse_tables(tables_data):
        """Parse extracted table data."""
        for table in tables_data:
            if len(table) < 2:
                continue
            
            header = table[0]
            header_str = ' '.join(str(h).lower() for h in header if h)
            
            for row in table[1:]:
                if not row or not any(row):
                    continue
                
                if 'department' in header_str:
                    if len(row) >= 2:
                        Department.objects.get_or_create(
                            code=str(row[1]),
                            defaults={'name': str(row[0])}
                        )
                
                elif 'lecturer' in header_str:
                    if len(row) >= 3:
                        try:
                            dept = Department.objects.get(code=str(row[2]))
                            Lecturer.objects.get_or_create(
                                employee_id=str(row[1]),
                                defaults={'name': str(row[0]), 'department': dept}
                            )
                        except Department.DoesNotExist:
                            pass
                
                elif 'room' in header_str:
                    if len(row) >= 2:
                        try:
                            capacity = int(str(row[1]))
                            Room.objects.get_or_create(
                                name=str(row[0]),
                                defaults={'capacity': capacity, 'room_type': 'CLASSROOM'}
                            )
                        except (ValueError, TypeError):
                            pass
                
                elif 'time' in header_str or 'slot' in header_str:
                    if len(row) >= 3:
                        try:
                            TimeSlot.objects.get_or_create(
                                day=str(row[0]),
                                start_time=str(row[1]),
                                end_time=str(row[2])
                            )
                        except Exception:
                            pass
                
                elif 'course' in header_str:
                    if len(row) >= 3:
                        try:
                            dept = Department.objects.get(code=str(row[2]))
                            Course.objects.get_or_create(
                                code=str(row[0]),
                                defaults={
                                    'name': str(row[1]),
                                    'department': dept,
                                    'year_of_study': 1,
                                    'study_mode': 'IN_PERSON',
                                    'class_size': 30,
                                    'hours_per_week': 1
                                }
                            )
                        except (Department.DoesNotExist, ValueError):
                            pass
    
    @staticmethod
    def _parse_text_sections(text_content):
        """Parse text-based sections (fallback method)."""
        sections = text_content.split('\n\n')
        
        for section in sections:
            section_lower = section.lower()
            
            if 'department' in section_lower:
                lines = section.strip().split('\n')
                for line in lines[1:]:
                    parts = line.split()
                    if len(parts) >= 2:
                        name = ' '.join(parts[:-1])
                        code = parts[-1]
                        if name and code:
                            Department.objects.get_or_create(
                                code=code,
                                defaults={'name': name}
                            )
        
        return True, "PDF data imported successfully"
    
    @staticmethod
    def parse_training_data_file(file):
        """
        Parse training data file and extract courses, lecturers, departments, rooms.
        
        Args:
            file: Uploaded file object
            
        Returns:
            tuple: (success: bool, message: str, data_summary: dict)
        """
        filename = file.name.lower()
        
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                return FileParser._parse_training_excel(file)
            elif filename.endswith('.csv'):
                return FileParser._parse_training_csv(file)
            else:
                return False, "Unsupported file type. Use .xlsx or .csv", {}
        except Exception as e:
            return False, f"Error parsing file: {str(e)}", {}
    
    @staticmethod
    def _parse_training_excel(file):
        """Parse Excel training data file."""
        
        try:
            wb = openpyxl.load_workbook(file)
            data_summary = {
                'schools': [],
                'departments': [],
                'courses': [],
                'lecturers': [],
                'rooms': []
            }
            
            # Parse Schools sheet
            if 'Schools' in wb.sheetnames:
                ws = wb['Schools']
                schools = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If school name exists
                        school, created = School.objects.get_or_create(
                            code=row[1] or row[0][:4].upper(),
                            defaults={
                                'name': row[0],
                                'description': row[2] if len(row) > 2 else ''
                            }
                        )
                        schools.append(school)
                        data_summary['schools'].append(school.id)
            
            # Parse Departments sheet
            if 'Departments' in wb.sheetnames:
                ws = wb['Departments']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If department name exists
                        school = None
                        if row[2]:  # If school code provided
                            try:
                                school = School.objects.get(code=row[2])
                            except School.DoesNotExist:
                                pass
                        
                        dept, created = Department.objects.get_or_create(
                            code=row[1] or row[0][:3].upper(),
                            defaults={
                                'name': row[0],
                                'school': school
                            }
                        )
                        data_summary['departments'].append(dept.id)
            
            # Parse Courses sheet
            if 'Courses' in wb.sheetnames:
                ws = wb['Courses']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If course code exists
                        try:
                            dept = Department.objects.get(code=row[2]) if row[2] else None
                        except Department.DoesNotExist:
                            dept = None
                        
                        course, created = Course.objects.get_or_create(
                            code=row[0],
                            defaults={
                                'name': row[1],
                                'department': dept,
                                'year_of_study': int(row[4]) if row[4] else 1,
                                'class_size': int(row[5]) if row[5] else 0
                            }
                        )
                        data_summary['courses'].append(course.id)
            
            # Parse Lecturers sheet
            if 'Lecturers' in wb.sheetnames:
                ws = wb['Lecturers']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If name exists
                        try:
                            dept = Department.objects.get(code=row[2]) if row[2] else None
                        except Department.DoesNotExist:
                            dept = None
                        
                        lecturer, created = Lecturer.objects.get_or_create(
                            employee_id=row[1] if row[1] else f"EMP-{row[0][:3]}",
                            defaults={
                                'name': row[0],
                                'department': dept,
                                'email': row[3] if len(row) > 3 else ''
                            }
                        )
                        data_summary['lecturers'].append(lecturer.id)
            
            # Parse Rooms sheet
            if 'Rooms' in wb.sheetnames:
                ws = wb['Rooms']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If room name exists
                        room, created = Room.objects.get_or_create(
                            name=row[0],
                            defaults={
                                'building': row[1] if len(row) > 1 else '',
                                'capacity': int(row[2]) if len(row) > 2 and row[2] else 30,
                                'room_type': row[3] if len(row) > 3 else 'Lecture_Hall'
                            }
                        )
                        data_summary['rooms'].append(room.id)
            
            message = f"Training data imported successfully: {len(data_summary['courses'])} courses, {len(data_summary['lecturers'])} lecturers, {len(data_summary['rooms'])} rooms"
            return True, message, data_summary
            
        except Exception as e:
            return False, f"Error processing Excel file: {str(e)}", {}
    
    @staticmethod
    def _parse_training_csv(file):
        """Parse CSV training data file."""
        import csv
        import io
        
        try:
            file_content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(file_content))
            
            data_summary = {
                'courses': [],
                'lecturers': [],
                'rooms': []
            }
            
            for row in reader:
                if row.get('course_code'):
                    course, created = Course.objects.get_or_create(
                        code=row.get('course_code'),
                        defaults={
                            'name': row.get('course_name', 'Unknown'),
                            'year_of_study': int(row.get('year', 1)),
                            'class_size': int(row.get('class_size', 0))
                        }
                    )
                    if created:
                        data_summary['courses'].append(course.id)
            
            message = f"CSV training data imported: {len(data_summary['courses'])} courses"
            return True, message, data_summary
            
        except Exception as e:
            return False, f"Error processing CSV file: {str(e)}", {}
    
    @staticmethod
    def parse_timetable_data_file(file):
        """
        Parse timetable data file for scheduling.
        
        Args:
            file: Uploaded file object
            
        Returns:
            tuple: (success: bool, message: str, data_summary: dict)
        """
        filename = file.name.lower()
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            return FileParser._parse_timetable_excel(file)
        else:
            return False, "Timetable data must be in Excel format", {}
    
    @staticmethod
    def _parse_timetable_excel(file):
        """Parse Excel timetable data and create timetable entries."""
        try:
            wb = openpyxl.load_workbook(file)
            data_summary = {'entries': 0, 'conflicts': []}
            
            if 'Timetable' in wb.sheetnames:
                ws = wb['Timetable']
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        course_code = row[0]
                        lecturer_id = row[1]
                        room_id = row[2]
                        day = row[3]
                        
                        course = Course.objects.get(code=course_code)
                        lecturer = Lecturer.objects.get(id=lecturer_id)
                        room = Room.objects.get(id=room_id)
                        
                        # Find time slot or create from data
                        time_slot = TimeSlot.objects.filter(day=day).first()
                        if not time_slot:
                            time_slot = TimeSlot.objects.create(
                                day=day,
                                start_time='09:00',
                                end_time='10:00'
                            )
                        
                        entry, created = TimetableEntry.objects.get_or_create(
                            course=course,
                            time_slot=time_slot,
                            defaults={
                                'lecturer': lecturer,
                                'room': room
                            }
                        )
                        
                        if created:
                            data_summary['entries'] += 1
                    except Exception as e:
                        data_summary['conflicts'].append(str(e))
            
            message = f"Timetable data processed: {data_summary['entries']} entries created"
            return True, message, data_summary
            
        except Exception as e:
            return False, f"Error processing timetable file: {str(e)}", {}
