"""
DRF viewsets for CRUD on all timetable models.
"""

from ortools.sat.python import cp_model
from .models import (
    User,
    Department,
    Course,
    Lecturer,
    Room,
    TimeSlot,
    LecturerAvailability,
    TimetableEntry,
)
from .serializers import (
    UserSerializer,
    DepartmentSerializer,
    CourseSerializer,
    LecturerSerializer,
    RoomSerializer,
    TimeSlotSerializer,
    LecturerAvailabilitySerializer,
    TimetableEntrySerializer,
)
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate
from django.views.decorators.http import require_http_methods
import openpyxl
import pdfplumber
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from datetime import datetime


class IsAdminUserOrReadOnly(permissions.BasePermission):
    """Allow full access to ADMIN users; read-only for VIEWER users."""

    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return (
            request.user.is_authenticated
            and hasattr(request.user, 'is_admin_user')
            and request.user.is_admin_user
        )


# ---------------------------------------------------------------------------
# Authentication endpoints
# ---------------------------------------------------------------------------

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register_user(request):
    """Register a new user (creates VIEWER role by default)."""
    try:
        username = request.data.get('username')
        password = request.data.get('password')
        email = request.data.get('email', '')
        first_name = request.data.get('first_name', '')
        last_name = request.data.get('last_name', '')
        
        if not username or not password:
            return Response(
                {'error': 'Username and password are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if User.objects.filter(username=username).exists():
            return Response(
                {'error': 'Username already exists.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=User.Role.VIEWER,  # Default role is VIEWER
        )
        
        serializer = UserSerializer(user)
        return Response(
            {'message': 'User registered successfully.', 'user': serializer.data},
            status=status.HTTP_201_CREATED
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_user(request):
    """Authenticate user and return user info."""
    try:
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response(
                {'error': 'Username and password are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = authenticate(username=username, password=password)
        
        if user is None:
            return Response(
                {'error': 'Invalid username or password.'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        serializer = UserSerializer(user)
        return Response(
            {'message': 'Login successful.', 'user': serializer.data},
            status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def current_user(request):
    """Get current authenticated user's information."""
    try:
        serializer = UserSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'PATCH'])
@permission_classes([permissions.IsAuthenticated])
def update_user_profile(request):
    """Update current user's profile information."""
    try:
        user = request.user
        
        # Update allowed fields
        if 'email' in request.data:
            user.email = request.data['email']
        if 'first_name' in request.data:
            user.first_name = request.data['first_name']
        if 'last_name' in request.data:
            user.last_name = request.data['last_name']
        
        # Password update (require current password)
        if 'new_password' in request.data:
            current_password = request.data.get('current_password')
            if not current_password or not user.check_password(current_password):
                return Response(
                    {'error': 'Current password is incorrect.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            user.set_password(request.data['new_password'])
        
        user.save()
        serializer = UserSerializer(user)
        return Response(
            {'message': 'Profile updated successfully.', 'user': serializer.data},
            status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def logout_user(request):
    """Logout user (for token-based auth, client deletes token)."""
    return Response(
        {'message': 'Logout successful. Please delete your token/session.'},
        status=status.HTTP_200_OK
    )


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    search_fields = ['name', 'code']


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related('department').all()
    serializer_class = CourseSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['department', 'year_of_study', 'study_mode']
    search_fields = ['name', 'code']


class LecturerViewSet(viewsets.ModelViewSet):
    queryset = Lecturer.objects.select_related('department').all()
    serializer_class = LecturerSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['department']
    search_fields = ['name', 'employee_id']


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['room_type', 'building']
    search_fields = ['name', 'building']


class TimeSlotViewSet(viewsets.ModelViewSet):
    queryset = TimeSlot.objects.all()
    serializer_class = TimeSlotSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['day']


class LecturerAvailabilityViewSet(viewsets.ModelViewSet):
    queryset = LecturerAvailability.objects.select_related(
        'lecturer', 'time_slot'
    ).all()
    serializer_class = LecturerAvailabilitySerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['lecturer', 'is_available']


class TimetableEntryViewSet(viewsets.ModelViewSet):
    queryset = TimetableEntry.objects.select_related(
        'course', 'lecturer', 'room', 'time_slot'
    ).all()
    serializer_class = TimetableEntrySerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUserOrReadOnly]
    filterset_fields = ['time_slot__day', 'is_locked', 'course', 'lecturer', 'room']


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminUserOrReadOnly])
def upload_data(request):
    """Upload and parse Excel or PDF file to create/update data."""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    file_name = file.name.lower()
    
    if file_name.endswith('.xlsx'):
        return parse_excel(file)
    elif file_name.endswith('.pdf'):
        return parse_pdf(file)
    else:
        return Response({'error': 'Unsupported file type'}, status=status.HTTP_400_BAD_REQUEST)


def parse_excel(file):
    """Parse Excel file with sheets: Departments, Courses, Lecturers, Rooms, TimeSlots."""
    try:
        wb = openpyxl.load_workbook(file)
        
        # Parse Departments
        if 'Departments' in wb.sheetnames:
            sheet = wb['Departments']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # name, code
                    Department.objects.get_or_create(name=row[0], code=row[1])
        
        # Parse Lecturers
        if 'Lecturers' in wb.sheetnames:
            sheet = wb['Lecturers']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # name, employee_id, department_code
                    dept = get_object_or_404(Department, code=row[2])
                    Lecturer.objects.get_or_create(
                        employee_id=row[1], defaults={
                            'name': row[0], 'department': dept
                        }
                    )
        
        # Parse Rooms
        if 'Rooms' in wb.sheetnames:
            sheet = wb['Rooms']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # name, capacity
                    Room.objects.get_or_create(
                        name=row[0], defaults={
                            'capacity': row[1] or 30, 'room_type': row[2] or 'CLASSROOM',
                            'building': row[3] or 'Main Building'
                        }
                    )
        
        # Parse TimeSlots
        if 'TimeSlots' in wb.sheetnames:
            sheet = wb['TimeSlots']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # day, start_time, end_time
                    TimeSlot.objects.get_or_create(
                        day=row[0], start_time=row[1], end_time=row[2]
                    )
        
        # Parse Courses (after departments are created)
        if 'Courses' in wb.sheetnames:
            sheet = wb['Courses']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # code, name, department_code
                    dept = get_object_or_404(Department, code=row[2])
                    Course.objects.get_or_create(
                        code=row[0], defaults={
                            'name': row[1], 'department': dept, 'year_of_study': row[3] or 1,
                            'study_mode': row[4] or 'IN_PERSON', 'class_size': row[5] or 30,
                            'hours_per_week': row[6] or 1
                        }
                    )
        
        return Response({'message': 'Excel data uploaded and parsed successfully'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': f'Error parsing Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


def parse_pdf(file):
    """Parse PDF file to extract tabular data for departments, courses, lecturers, rooms, timeslots."""
    try:
        with pdfplumber.open(file) as pdf:
            text_content = ""
            tables_data = []
            
            for page in pdf.pages:
                # Extract text
                text_content += page.extract_text() + "\n"
                
                # Extract tables
                tables = page.extract_tables()
                for table in tables:
                    if table and len(table) > 1:  # Has header and data
                        tables_data.append(table)
            
            # Try to identify and parse different sections
            sections = text_content.split('\n\n')
            
            for section in sections:
                section_lower = section.lower()
                
                # Parse Departments section
                if 'department' in section_lower:
                    lines = section.strip().split('\n')
                    for line in lines[1:]:  # Skip header
                        parts = line.split()
                        if len(parts) >= 2:
                            name = ' '.join(parts[:-1])
                            code = parts[-1]
                            if name and code:
                                Department.objects.get_or_create(name=name, code=code)
                
                # Parse Lecturers section
                elif 'lecturer' in section_lower:
                    lines = section.strip().split('\n')
                    for line in lines[1:]:  # Skip header
                        parts = line.split()
                        if len(parts) >= 3:
                            name = ' '.join(parts[:-2])
                            employee_id = parts[-2]
                            dept_code = parts[-1]
                            try:
                                dept = Department.objects.get(code=dept_code)
                                Lecturer.objects.get_or_create(
                                    employee_id=employee_id, 
                                    defaults={'name': name, 'department': dept}
                                )
                            except Department.DoesNotExist:
                                continue
                
                # Parse Rooms section
                elif 'room' in section_lower:
                    lines = section.strip().split('\n')
                    for line in lines[1:]:  # Skip header
                        parts = line.split()
                        if len(parts) >= 2:
                            name = parts[0]
                            try:
                                capacity = int(parts[1])
                                Room.objects.get_or_create(
                                    name=name, 
                                    defaults={'capacity': capacity, 'room_type': 'CLASSROOM'}
                                )
                            except ValueError:
                                continue
                
                # Parse TimeSlots section
                elif 'time' in section_lower or 'slot' in section_lower:
                    lines = section.strip().split('\n')
                    for line in lines[1:]:  # Skip header
                        parts = line.split()
                        if len(parts) >= 3:
                            day = parts[0]
                            start_time = parts[1]
                            end_time = parts[2]
                            TimeSlot.objects.get_or_create(
                                day=day, start_time=start_time, end_time=end_time
                            )
                
                # Parse Courses section
                elif 'course' in section_lower:
                    lines = section.strip().split('\n')
                    for line in lines[1:]:  # Skip header
                        parts = line.split()
                        if len(parts) >= 3:
                            code = parts[0]
                            name = ' '.join(parts[1:-1])
                            dept_code = parts[-1]
                            try:
                                dept = Department.objects.get(code=dept_code)
                                Course.objects.get_or_create(
                                    code=code,
                                    defaults={
                                        'name': name, 'department': dept, 'year_of_study': 1,
                                        'study_mode': 'IN_PERSON', 'class_size': 30, 'hours_per_week': 1
                                    }
                                )
                            except Department.DoesNotExist:
                                continue
            
            # Also try to parse tables if found
            for table in tables_data:
                if len(table) < 2:
                    continue
                    
                header = table[0]
                header_str = ' '.join(str(h) for h in header if h).lower()
                
                for row in table[1:]:
                    if not row or not any(row):
                        continue
                        
                    if 'department' in header_str:
                        if len(row) >= 2:
                            Department.objects.get_or_create(name=str(row[0]), code=str(row[1]))
                    
                    elif 'lecturer' in header_str:
                        if len(row) >= 3:
                            try:
                                dept = Department.objects.get(code=str(row[2]))
                                Lecturer.objects.get_or_create(
                                    employee_id=str(row[1]), 
                                    defaults={'name': str(row[0]), 'department': dept}
                                )
                            except Department.DoesNotExist:
                                continue
                    
                    elif 'room' in header_str:
                        if len(row) >= 2:
                            try:
                                capacity = int(str(row[1]))
                                Room.objects.get_or_create(
                                    name=str(row[0]), 
                                    defaults={'capacity': capacity, 'room_type': 'CLASSROOM'}
                                )
                            except ValueError:
                                continue
                    
                    elif 'time' in header_str or 'slot' in header_str:
                        if len(row) >= 3:
                            TimeSlot.objects.get_or_create(
                                day=str(row[0]), start_time=str(row[1]), end_time=str(row[2])
                            )
                    
                    elif 'course' in header_str:
                        if len(row) >= 3:
                            try:
                                dept = Department.objects.get(code=str(row[2]))
                                Course.objects.get_or_create(
                                    code=str(row[0]),
                                    defaults={
                                        'name': str(row[1]), 'department': dept, 'year_of_study': 1,
                                        'study_mode': 'IN_PERSON', 'class_size': 30, 'hours_per_week': 1
                                    }
                                )
                            except Department.DoesNotExist:
                                continue
        
        return Response({'message': 'PDF data uploaded and parsed successfully'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': f'Error parsing PDF: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_excel_template(request):
    """Generate and return a sample Excel template for data upload."""
    try:
        wb = openpyxl.Workbook()
        
        # Departments sheet
        ws_dept = wb.active
        ws_dept.title = "Departments"
        ws_dept['A1'] = "Department Name"
        ws_dept['B1'] = "Department Code"
        ws_dept['A2'] = "Computer Science"
        ws_dept['B2'] = "CS"
        ws_dept['A3'] = "Mathematics"
        ws_dept['B3'] = "MATH"
        
        # Lecturers sheet
        ws_lect = wb.create_sheet("Lecturers")
        ws_lect['A1'] = "Lecturer Name"
        ws_lect['B1'] = "Employee ID"
        ws_lect['C1'] = "Department Code"
        ws_lect['A2'] = "Dr. John Smith"
        ws_lect['B2'] = "LEC001"
        ws_lect['C2'] = "CS"
        
        # Rooms sheet
        ws_rooms = wb.create_sheet("Rooms")
        ws_rooms['A1'] = "Room Name"
        ws_rooms['B1'] = "Capacity"
        ws_rooms['C1'] = "Room Type"
        ws_rooms['D1'] = "Building"
        ws_rooms['A2'] = "Room 101"
        ws_rooms['B2'] = 50
        ws_rooms['C2'] = "CLASSROOM"
        ws_rooms['D2'] = "Main Building"
        
        # TimeSlots sheet
        ws_times = wb.create_sheet("TimeSlots")
        ws_times['A1'] = "Day"
        ws_times['B1'] = "Start Time"
        ws_times['C1'] = "End Time"
        ws_times['A2'] = "Monday"
        ws_times['B2'] = "09:00"
        ws_times['C2'] = "10:00"
        ws_times['A3'] = "Monday"
        ws_times['B3'] = "10:00"
        ws_times['C3'] = "11:00"
        
        # Courses sheet
        ws_courses = wb.create_sheet("Courses")
        ws_courses['A1'] = "Course Code"
        ws_courses['B1'] = "Course Name"
        ws_courses['C1'] = "Department Code"
        ws_courses['D1'] = "Year of Study"
        ws_courses['E1'] = "Study Mode"
        ws_courses['F1'] = "Class Size"
        ws_courses['G1'] = "Hours per Week"
        ws_courses['A2'] = "CS101"
        ws_courses['B2'] = "Introduction to Programming"
        ws_courses['C2'] = "CS"
        ws_courses['D2'] = 1
        ws_courses['E2'] = "IN_PERSON"
        ws_courses['F2'] = 30
        ws_courses['G2'] = 2
        
        # Save to memory
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="timetable_template.xlsx"'
        return response
        
    except Exception as e:
        return Response({'error': f'Error generating template: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def course_timetable(request, course_id):
    """Get timetable entries for a specific course."""
    entries = TimetableEntry.objects.filter(course_id=course_id).select_related('time_slot', 'room', 'lecturer')
    serializer = TimetableEntrySerializer(entries, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def lecturer_timetable(request, lecturer_id):
    """Get timetable entries for a specific lecturer."""
    entries = TimetableEntry.objects.filter(lecturer_id=lecturer_id).select_related('time_slot', 'room', 'course')
    serializer = TimetableEntrySerializer(entries, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def room_timetable(request, room_id):
    """Get timetable entries for a specific room."""
    entries = TimetableEntry.objects.filter(room_id=room_id).select_related('time_slot', 'course', 'lecturer')
    serializer = TimetableEntrySerializer(entries, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_timetable_pdf(request):
    """Export timetable as PDF with optional filtering by lecturer/room/course."""
    try:
        lecturer_id = request.query_params.get('lecturer_id')
        room_id = request.query_params.get('room_id')
        course_id = request.query_params.get('course_id')
        
        # Build query
        query = TimetableEntry.objects.select_related('course', 'lecturer', 'room', 'time_slot')
        
        if lecturer_id:
            query = query.filter(lecturer_id=lecturer_id)
        if room_id:
            query = query.filter(room_id=room_id)
        if course_id:
            query = query.filter(course_id=course_id)
        
        entries = list(query.order_by('time_slot__day', 'time_slot__start_time'))
        
        if not entries:
            return Response(
                {'error': 'No timetable entries found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
        )
        
        if lecturer_id:
            lecturer = get_object_or_404(Lecturer, id=lecturer_id)
            title = f"Timetable for {lecturer.name}"
        elif room_id:
            room = get_object_or_404(Room, id=room_id)
            title = f"Timetable for {room.name}"
        elif course_id:
            course = get_object_or_404(Course, id=course_id)
            title = f"Timetable for {course.code}: {course.name}"
        else:
            title = "Complete Timetable"
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Generated date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
        )
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Prepare table data
        data = [['Day', 'Time', 'Course', 'Lecturer', 'Room', 'Class Size']]
        
        for entry in entries:
            day = entry.time_slot.get_day_display()
            time = f"{entry.time_slot.start_time.strftime('%H:%M')} - {entry.time_slot.end_time.strftime('%H:%M')}"
            course = f"{entry.course.code}"
            lecturer = entry.lecturer.name
            room = entry.room.name
            class_size = entry.course.class_size
            
            data.append([day, time, course, lecturer, room, str(class_size)])
        
        # Create table
        table = Table(data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsAdminUserOrReadOnly])
def generate_timetable(request):
    """Generate timetable using AI/ML optimization with constraint programming."""
    try:
        # Clear existing timetable
        TimetableEntry.objects.all().delete()
        
        # Get all data
        courses = list(Course.objects.all())
        lecturers = list(Lecturer.objects.all())
        rooms = list(Room.objects.all())
        time_slots = list(TimeSlot.objects.all())
        availabilities = list(LecturerAvailability.objects.all())
        
        if not courses or not lecturers or not rooms or not time_slots:
            return Response({'error': 'Insufficient data for timetable generation'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Use OR-Tools for constraint programming
        model = cp_model.CpModel()
        
        # Create variables: course_time_room[course][time][room]
        course_time_room = {}
        for c_idx, course in enumerate(courses):
            course_time_room[c_idx] = {}
            for t_idx, time_slot in enumerate(time_slots):
                course_time_room[c_idx][t_idx] = {}
                for r_idx, room in enumerate(rooms):
                    course_time_room[c_idx][t_idx][r_idx] = model.NewBoolVar(f'course_{c_idx}_time_{t_idx}_room_{r_idx}')
        
        # Each course must be scheduled exactly once
        for c_idx in range(len(courses)):
            model.Add(sum(course_time_room[c_idx][t_idx][r_idx] 
                         for t_idx in range(len(time_slots)) 
                         for r_idx in range(len(rooms))) == 1)
        
        # No two courses in same time slot and room
        for t_idx in range(len(time_slots)):
            for r_idx in range(len(rooms)):
                model.Add(sum(course_time_room[c_idx][t_idx][r_idx] 
                             for c_idx in range(len(courses))) <= 1)
        
        # Room capacity constraints
        for c_idx, course in enumerate(courses):
            for t_idx in range(len(time_slots)):
                for r_idx, room in enumerate(rooms):
                    if room.capacity < course.class_size:
                        model.Add(course_time_room[c_idx][t_idx][r_idx] == 0)
        
        # Lecturer availability and assignment
        lecturer_course = {}
        for c_idx, course in enumerate(courses):
            lecturer = next((l for l in lecturers if l.department == course.department), lecturers[0])
            lecturer_course[c_idx] = lecturer
            
            for t_idx, time_slot in enumerate(time_slots):
                # Check availability
                available = not any(a.time_slot == time_slot and a.lecturer == lecturer and not a.available 
                                  for a in availabilities)
                if not available:
                    for r_idx in range(len(rooms)):
                        model.Add(course_time_room[c_idx][t_idx][r_idx] == 0)
        
        # Lecturer can't teach multiple courses at same time
        for l_idx, lecturer in enumerate(lecturers):
            courses_for_lecturer = [c_idx for c_idx, c in enumerate(courses) if lecturer_course[c_idx] == lecturer]
            for t_idx in range(len(time_slots)):
                model.Add(sum(course_time_room[c_idx][t_idx][r_idx] 
                             for c_idx in courses_for_lecturer 
                             for r_idx in range(len(rooms))) <= 1)
        
        # Solve the model
        solver = cp_model.CpSolver()
        status = solver.Solve(model)
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            # Create timetable entries
            for c_idx, course in enumerate(courses):
                for t_idx, time_slot in enumerate(time_slots):
                    for r_idx, room in enumerate(rooms):
                        if solver.Value(course_time_room[c_idx][t_idx][r_idx]):
                            lecturer = lecturer_course[c_idx]
                            TimetableEntry.objects.create(
                                course=course,
                                lecturer=lecturer,
                                room=room,
                                time_slot=time_slot
                            )
            
            return Response({'message': 'Timetable generated successfully using AI constraint optimization'}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'No feasible timetable found'}, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
